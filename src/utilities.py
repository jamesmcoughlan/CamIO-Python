#utilities.py -- utility functions for CamIO
#(c) James Coughlan, Smith-Kettlewell Eye Research Institute

#uses dictionary styles instead of lists;
#adds ability to save object.

import numpy as np
import cv2
from openpyxl import Workbook, load_workbook

def list2dict(L): #convert list to dict with keys ranging 1, 2, 3, ...
	return {(k+1):L[k] for k in range(len(L))}

def dict2list(d): #convert dict with keys ranging 1, 2, 3, ... to list
	return [d[k+1] for k in range(len(d))]
	
def load_object(object_fname): #load CamIO object model from Excel spreadsheet	
	wb = load_workbook(filename = object_fname)
	sheet_list = wb.sheetnames
	sheet = wb[sheet_list[0]]
	
	#count how many hotspots there are:
	r = 1
	while True:
		#print('cell:',sheet.cell(row=r,column=1).value)
		if sheet.cell(row=r,column=1).value == None:
			break
		r += 1
	#now r is (the number of rows) + 1
	number_hotspots = r - 5

	board_parameters = []
	for k in range(4):
		board_parameters.append(sheet.cell(row=3,column=2+k).value)
	board_parameters[2] += 0. #revised 11/13/17 to handle case where spreadsheet contains integer entries
	board_parameters[3] += 0. #revised 11/13/17 to handle case where spreadsheet contains integer entries
	
	hotspots, labels, labels_secondary = [], [], []
	for h in range(number_hotspots):
		X,Y,Z = sheet.cell(row=5+h,column=4).value,sheet.cell(row=5+h,column=5).value,sheet.cell(row=5+h,column=6).value
		hotspots.append([X+0.,Y+0.,Z+0.]) #revised 11/13/17 to handle case where spreadsheet contains integer entries
		labels.append(sheet.cell(row=5+h,column=2).value)
		labels_secondary.append(sheet.cell(row=5+h,column=3).value)
					
	return wb, sheet, board_parameters, list2dict(hotspots), list2dict(labels), list2dict(labels_secondary)	
	
def save_object(wb, sheet, object_fname, hotspots3D, labels, secondary_labels): #save CamIO object model to Excel file
	#wb = Workbook()
	#sheet = wb.active
	
	#fill in spreadsheet template items:
	sheet.cell(row=4,column=1).value = 'Hotspot #'; sheet.cell(row=4,column=2).value = 'Label'; sheet.cell(row=4,column=3).value = 'Secondary label'
	sheet.cell(row=4,column=4).value = 'X'; sheet.cell(row=4,column=5).value = 'Y'; sheet.cell(row=4,column=6).value = 'Z'
		
	for k in hotspots3D:
		sheet.cell(row=4+k,column=1).value = k
		sheet.cell(row=4+k,column=2).value = labels[k] #don't deal with secondary labels
		for ind in range(3): #x,y,z indices
			sheet.cell(row=4+k,column=4+ind).value = hotspots3D[k][ind]

	wb.save(object_fname)
	
def get_board_pose(gray, this_dictionary, these_parameters, board, mtx, dist):
	corners, ids, rejectedImgPoints = cv2.aruco.detectMarkers(gray, dictionary=this_dictionary, parameters=these_parameters)
	#print 'blah:',corners
	if len(corners)>0: #otherwise the program can crash if it doesn't see any markers
		retval, charucoCorners, charucoIds = cv2.aruco.interpolateCornersCharuco(corners, ids, gray, board) #[, charucoCorners[, charucoIds[, cameraMatrix[, distCoeffs]]]]) -> retval, charucoCorners, charucoIds
		retval, rvec, tvec = cv2.aruco.estimatePoseCharucoBoard(charucoCorners, charucoIds, board, mtx, dist) #[, rvec[, tvec]]) -> retval, rvec, tvec
		if retval:
			return [rvec, tvec]
		else:
			return None

#################################################

def convert_coors_from_camera_to_marker(P,rvec,tvec):
	#P is the 3D point (in camera coordinates) to be converted to marker/board (such as marker or charuco board) coordinate system.
	#Coordinate system is specified by rvec,tvec, returned from a marker/board detection.
	#Note that rvec,tvec are numpy arrays, not matrices
	
	#find probe location in marker/board coordinates:
	R, _ = cv2.Rodrigues(rvec)
	R = np.matrix(R)
	P = np.matrix(P).transpose() #create a column vector
	t = np.matrix(tvec).transpose() #create a column vector
	P2 = np.linalg.inv(R)*(P - t)
	X,Y,Z = P2[0,0],P2[1,0],P2[2,0]
	
	return X,Y,Z

def convert_coors_from_marker_to_camera(P,rvec,tvec):
	#Opposite of above function.
	#Coordinate system is specified by rvec,tvec, returned from a marker/board detection.
	#Note that rvec,tvec are numpy arrays, not matrices.
	
	R, _ = cv2.Rodrigues(rvec)
	R = np.matrix(R)
	P = np.matrix(P).transpose() #create a column vector
	t = np.matrix(tvec).transpose() #create a column vector
	P2 = R*P + t
	X,Y,Z = P2[0,0],P2[1,0],P2[2,0]
	
	return X,Y,Z
	

#geometry

def get_rvec_tvec(img, mtx, dist, board,arucodict,arucodetectparam): #get rotation and translation of charuco board in an RGB image
	#also return grayscale version of image and a flag that is 1 if charuco pose has been estimated or -1 otherwise
	gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
	corners, ids, rejectedImgPoints = cv2.aruco.detectMarkers(gray, dictionary=arucodict, parameters=arucodetectparam)
	flag = -1 #not OK; if we get a valid charuco pose then we'll change to 1
	if len(corners)>0: #otherwise the program can crash if it doesn't see any markers
		retval, charucoCorners, charucoIds = cv2.aruco.interpolateCornersCharuco(corners, ids, gray, board) #[, charucoCorners[, charucoIds[, cameraMatrix[, distCoeffs]]]]) -> retval, charucoCorners, charucoIds
		retval, rvec, tvec = cv2.aruco.estimatePoseCharucoBoard(charucoCorners, charucoIds, board, mtx, dist) #[, rvec[, tvec]]) -> retval, rvec, tvec
		if retval:
			flag = 1
		
	return rvec, tvec, gray, flag
	
def dist2hotspots2D(together, hotspots, rvec, tvec, mtx, dist): #2D (image) dist from (row,col) to all hotspot locations in image
	row,col = together
	dlist = []
	for k, (X, Y, Z) in enumerate(hotspots):	
		objp2 = np.array(((X,Y,Z),)) #should be n x 3 where n = 1 in this case
		imgpoints, _ = cv2.projectPoints(objp2, rvec[:,0], tvec[:,0], mtx, dist)
		point = np.array([imgpoints[0][0][0],imgpoints[0][0][1]])
		d = cv2.norm(point - np.array((col,row))) #note (col, row) vs. (row, col)
		dlist.append(d)
		
	return dlist
	
def dist2hotspots3D(pt3D, hotspots): #3D dist from pt3D to all 3D hotspot locations
	dlist = []
	for (X, Y, Z) in dict2list(hotspots):	
		d = cv2.norm(pt3D - np.array((X,Y,Z)))
		dlist.append(d)
		
	return dlist
	
#more geometry
def find_closest_normal(rvec, normal): #find the direction in the coordinate system specified by rvec that is best aligned to normal vector
	R,_ = cv2.Rodrigues(rvec)
	a,b,c,a2,b2,c2 = R[:,0],R[:,1],R[:,2],-R[:,0],-R[:,1],-R[:,2]
	directions = [a,b,c,a2,b2,c2]
	L = [a@normal,b@normal,c@normal,a2@normal,b2@normal,c2@normal]
	index = np.argmax(L)	
	return directions[index]
	
def get_ground_rvec_tvec2(stylus_object1, stylus_object2): #get the rvec, tvec corresponding to the ground plane from two views of stylus flat on table
	#Approach:
	#1) Get normal direction from rvecs (both views should have similar rvec so average them together). 
	#2) In effect, follow the same procedure before (see OLD below), but define the normal as explained above instead of using vectors on the plane.
	#3) Up direction is determined by constraint that camera line of sight points more down than up.
	#4) Desired ground plane is just a bit lower (half width of dowel) than center plane.
	#OLD approach:
	#1) Use origin and tip of one view and origin of second view to define plane of center points, which we call the center plane.
	#2) Up direction is determined by constraint that camera line of sight points more down than up.
	#3) Desired ground plane is just a bit lower (half width of dowel) than center plane.
	
	origin1 = stylus_object1.tvec
	R1,_ = cv2.Rodrigues(stylus_object1.rvec)
	tip1 = (R1@stylus_object1.offset) + origin1
	origin2 = stylus_object2.tvec

	#get approximate normal first:
	v1, v2 = tip1-origin1, origin2-origin1
	normal = np.cross(v1, v2) #should do something if normal is very small or 0!!!
	if normal[2] > 0: #make sure normal points up (out of table, approx. opposite camera line of sight), not down
		normal *= -1
	normal /= cv2.norm(normal) #make it unit norm
	
	#then refine the normal direction from the two rvec vectors:
	normal1 = find_closest_normal(stylus_object1.rvec, normal)
	normal2 = find_closest_normal(stylus_object2.rvec, normal)
	normal_avg = (normal1+normal2)/2
	normal_avg /= cv2.norm(normal_avg)
	
	#need two vectors perpendicular to normal_avg:
	x_avg = np.cross(normal_avg, v1)
	x_avg /= cv2.norm(x_avg)
	y_avg = np.cross(normal_avg, x_avg)
	y_avg /= cv2.norm(y_avg)
	
	R_avg = np.zeros((3,3),float)
	R_avg[:,0] = x_avg
	R_avg[:,1] = y_avg
	R_avg[:,2] = normal_avg
	rvec, _ = cv2.Rodrigues(R_avg)
	
	tvec = origin1 - (stylus_object.b/2)*normal_avg #subtract half the width of the stylus
	
	return np.squeeze(rvec), np.squeeze(tvec)

def assemble_transformation(rvec, tvec): #assemble 4x4 transformation matrix
	R,_ = cv2.Rodrigues(rvec)
	T = np.zeros((4,4),float)
	T[0:3,0:3] = R + 0
	T[0:3,3] = tvec
	T[3,3] = 1.0
	return T #return as numpy array

def append_1(v): #append a 1.0 at end of numpy vec
	return np.concatenate((v,np.array([1.])))
	
def get_transformation_3dof(Tcm, anchor1, anchor2, hotspot1, hotspot2):
	#get transformation assuming only rotation and translation on ground plane
	anchor1c, anchor2c = append_1(anchor1), append_1(anchor2)
	anchor1m, anchor2m = Tcm@anchor1c, Tcm@anchor2c #now in marker coordinates
	anchor1_xy, anchor2_xy = anchor1m[0:2], anchor2m[0:2] #remove the z and trailing 1 entries
	hotspot1_xy, hotspot2_xy = hotspot1[0:2], hotspot2[0:2]
	
	d_hotspot = hotspot2_xy - hotspot1_xy
	d_anchor = anchor2_xy - anchor1_xy
	norm_h = d_hotspot/cv2.norm(d_hotspot)
	perp_h = -np.array([norm_h[1],-norm_h[0]]) #sign?
	norm_d = d_anchor/cv2.norm(d_anchor)
	c=np.dot(norm_h,norm_d) 
	s=np.dot(perp_h,norm_d)
	rot = atan2(s,c)
	
	#trans = hotspot1_xy - anchor1_xy
	
	R_2x2 = np.array(([c, s], [-s, c]))
	trans = hotspot1_xy - R_2x2@anchor1_xy
	
	T = np.zeros((4,4),float)
	#T[0,0],T[0,1] = cos(rot), sin(rot)
	#T[1,0],T[1,1] = -sin(rot), cos(rot)
	T[0,0],T[0,1] = c,s
	T[1,0],T[1,1] = -s,c
	T[2,2] = 1.0
	T[0,3],T[1,3] = trans[0],trans[1]
	T[3,3] = 1.0
	
	return T
	
def determine_poses(plane_pose, anchor1, anchor2, hotspot1, hotspot2):
	#c = camera frame
	#m = marker frame
	#a = annotation frame
	#Tij = transformation matrix from frame i to frame j
	Tmc = assemble_transformation(plane_pose[0], plane_pose[1])
	Tcm = inv(Tmc)
	Tma = get_transformation_3dof(Tcm, anchor1, anchor2, hotspot1, hotspot2)
	Tca = Tma@Tcm
	Tac = inv(Tca)

	return Tmc, Tcm, Tma, Tca, Tac
			
def convert4x4_to_rvec_tvec(T):
	R = T[0:3,0:3]
	rvec, _ = cv2.Rodrigues(R)
	tvec = T[0:3,3]
	return np.squeeze(rvec), np.squeeze(tvec)
	
##################################################

